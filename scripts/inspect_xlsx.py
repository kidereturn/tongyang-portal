# -*- coding: utf-8 -*-
"""두 xlsx 파일의 sheet 구조와 셀 내용을 한눈에 출력"""
import sys, os
try: sys.stdout.reconfigure(encoding='utf-8')
except: pass

from openpyxl import load_workbook

paths = [
    r"C:\Users\tyinc\Downloads\바이브코딩 개발사례 - 2차강의\5. (표로 요약) 바이브코딩 강의.xlsx",
    r"C:\Users\tyinc\Downloads\바이브코딩 강의내용 공유(그룹 부동산관리 시스템_강사 권오윤)_26.03.24\1. 강의요약표 (개발노하우 별도 요약).xlsx",
]

for p in paths:
    print('='*100)
    print(f"FILE: {os.path.basename(p)}")
    print('='*100)
    if not os.path.exists(p):
        print(f"  NOT FOUND")
        continue
    wb = load_workbook(p, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n[Sheet: {sn}]  rows={ws.max_row} cols={ws.max_column}")
        max_r = min(ws.max_row, 80)
        max_c = min(ws.max_column, 12)
        for r in range(1, max_r + 1):
            row = []
            for c in range(1, max_c + 1):
                v = ws.cell(r, c).value
                s = '' if v is None else str(v).replace('\n', ' / ').strip()
                if len(s) > 60: s = s[:60] + '…'
                row.append(s)
            print(f"  R{r:02d} | " + " | ".join(row))
        print()
