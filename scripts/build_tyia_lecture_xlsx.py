# -*- coding: utf-8 -*-
"""
TYIA(내부회계 포털) 강의요약표 xlsx 생성
- 양식: 권오윤 차장 강의 요약표 양식 그대로
- Sheet 1: 카테고리 형식 (알아두면 좋은 용어, 개발환경, 인증·인프라, AI, 보안, 주요 기능)
- Sheet 2: 표 형식 (항목 / 용도 / 기술 / 세부 사양 / 보안 / 비고)
- Sheet 3: 개발 노하우 + 메뉴 범주별 요약표
"""
import os, sys
try: sys.stdout.reconfigure(encoding='utf-8')
except: pass

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'docs', 'TYIA_강의요약표.xlsx')

# ──────────────────────────────────────────
# 스타일
# ──────────────────────────────────────────
TITLE_FONT  = Font(name='맑은 고딕', size=14, bold=True, color='0F1E36')
H_FONT      = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
SUB_FONT    = Font(name='맑은 고딕', size=11, bold=True, color='3182F6')
CELL_FONT   = Font(name='맑은 고딕', size=10)
NOTE_FONT   = Font(name='맑은 고딕', size=10, italic=True, color='8B95A1')

H_FILL      = PatternFill('solid', fgColor='3182F6')
SUB_FILL    = PatternFill('solid', fgColor='EEF4FE')
ALT_FILL    = PatternFill('solid', fgColor='F9FAFB')

THIN = Side(style='thin', color='E5E8EB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

wb = Workbook()

# ════════════════════════════════════════════
# Sheet 1 — 카테고리 형식
# ════════════════════════════════════════════
ws = wb.active
ws.title = '내용-1(카테고리)'

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 4
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 4
ws.column_dimensions['F'].width = 4
ws.column_dimensions['G'].width = 4

r = 1
# 제목
ws.cell(r, 1, '■ TYIA 내부회계 포털 — 시스템 강의 요약 (2026.04.28)').font = TITLE_FONT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 2

# 카테고리 데이터
def write_section(title, items):
    global r
    c = ws.cell(r, 1, title)
    c.font = SUB_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    for label, desc in items:
        ws.cell(r, 2, label).font = CELL_FONT
        ws.cell(r, 2).alignment = LEFT
        ws.cell(r, 2).border = BORDER
        cv = ws.cell(r, 4, desc)
        cv.font = CELL_FONT
        cv.alignment = LEFT
        cv.border = BORDER
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1

write_section('1. 알아두면 좋은 용어', [
    ('TYIA',          '동양 내부회계 포털 시스템 이름 (Tongyang Internal Accounting)'),
    ('SPA',           'Single Page Application — 페이지 전환 없이 빠르게 동작하는 웹앱'),
    ('RLS',           'Row Level Security — 같은 테이블이라도 사용자별로 볼 수 있는 행이 다르게 적용되는 보안 정책'),
    ('OWNER',         '담당자 — 본인 통제활동의 증빙을 업로드·상신하는 역할 (442명)'),
    ('CONTROLLER',    '승인자 — 담당자가 상신한 증빙을 검토·승인/반려하는 역할 (33명)'),
    ('ADMIN',         '관리자 — 전사 관리·모니터링·검토 메모를 작성하는 역할 (3명, 내부회계팀)'),
    ('통제활동',       '내부회계관리제도에서 정의한 부서별 점검 항목 (예: 자금 결재, 매출 인식 등)'),
    ('모집단',         '통제활동의 점검 대상이 되는 거래/문서 목록 (예: 11월 자금 결재 100건)'),
    ('증빙',           '통제 수행을 입증하는 파일 (PDF·Excel — 결재 화면, 명세서 등)'),
    ('결재상신',       '담당자가 모든 모집단 증빙 업로드 완료 후 승인자에게 검토 요청을 보내는 동작'),
])

write_section('2. 개발 환경', [
    ('Node.js',       '20+ — 프론트엔드 빌드용 런타임'),
    ('npm',           '10+ — 패키지 매니저'),
    ('Vite',          '8.0 — 개발 서버 + 빌드 도구 (HMR 핫 리로드 지원)'),
    ('TypeScript',    '6.0 — JavaScript 에 타입 안정성을 더한 언어'),
    ('Viewport',      '데스크톱 1280~1920px 우선, 모바일은 추후 대응'),
])

write_section('3. 주요 패키지', [
    ('React',         '19.2 — UI 프레임워크 (Toss·구글이 사용하는 동일 라이브러리)'),
    ('Tailwind CSS',  '3.4 — 유틸리티 기반 CSS (사이트 디자인 톤 일관성)'),
    ('Recharts',      '3.8 — KPI 대시보드의 차트 (Bar/Pie/Trend)'),
    ('Lucide React',  '1.7 — 아이콘 라이브러리'),
    ('mammoth',       '1.12 — DOCX/HWP 파일 미리보기 변환'),
    ('pdfjs-dist',    '5.6 — PDF 미리보기 렌더링'),
    ('xlsx',          '0.18 — Excel 다운로드 (모집단·증빙목록)'),
    ('jszip',         '3.10 — 증빙 일괄 ZIP 다운로드'),
    ('react-router',  '7.14 — 페이지 라우팅'),
])

write_section('4. 인프라 (외부 API · 서비스)', [
    ('Supabase',      'PostgreSQL DB + 인증 + 파일 스토리지 (월 25달러 Pro 플랜)'),
    ('Vercel',        '프론트엔드 배포 + CDN (자동 배포, 무료 플랜)'),
    ('GitHub',        '소스 코드 버전 관리'),
    ('DART OpenAPI',  '회사소식 페이지 — 동양 공시·재무제표 자동 수집'),
    ('네이버 금융',    '주가 정보 + 뉴스 RSS'),
    ('Google News RSS', '레미콘·건설·플랜트 뉴스 자동 수집'),
])

write_section('5. AI 활용', [
    ('Claude (4.5/4.6)', '주력 코딩 도구 — 사이트 전체 개발에 사용'),
    ('Anthropic API',    'AI 챗봇 백엔드 (사용자 질의응답)'),
    ('OpenAI ChatGPT',   '아이디어 검토·문서 초안 보조'),
    ('AI 챗봇',          '사이트 내 /chatbot — 내부회계 관련 질의응답 24시간'),
])

write_section('6. 보안 정책', [
    ('RLS',                 'Row Level Security — 사용자별로 본인 데이터만 SELECT 가능 (DB 레벨)'),
    ('HTTPS 강제',          'Vercel 자동 SSL — 모든 통신 TLS 1.3 암호화'),
    ('비밀번호 해싱',        'Supabase Auth (bcrypt) — 평문 저장 안 됨'),
    ('가입승인',             '관리자 승인 후 활성화 (is_active = true)'),
    ('자동 로그아웃',         '24시간 inactive 시 세션 만료'),
    ('백업',                 'Supabase 자동 일일백업 7일 보관 + 주1회 수동 export'),
    ('SECURITY DEFINER RPC', '권한 우회가 필요한 작업은 SECURITY DEFINER 함수로 한정'),
])

write_section('7. 주요 기능', [
    ('증빙관리',     '담당자가 모집단별 증빙을 업로드, 자동 중간저장, 결재상신'),
    ('승인/반려',    '승인자가 검토 후 즉시 처리, 반려 시 사유 입력하면 담당자 알림'),
    ('관리자 메모',  '관리자가 검토결과 + 메모 작성, 담당자·승인자에게 표시'),
    ('수정제출',     '관리자가 승인된 항목도 재상신 요청 가능 (모든 승인 자동 취소)'),
    ('KPI 대시보드', '부서별 점수 + 등급(A~F), 월별 달성률 추이'),
    ('회사소식',     'DART 공시 + 동양 주가 + 뉴스 자동 수집 (4개 카테고리)'),
    ('빙고게임',     '매일 25문제, 1줄 완성/오답 시 도전 1회 소모, 일 최대 3회'),
    ('강좌',         '내부회계 교육 영상 + 수료 퀴즈 + 익명 질문 가능'),
    ('웹툰',         '내부회계 관련 짧은 만화 + 익명 댓글'),
    ('Tell me',      '회사에 대한 의견·불편·건의 익명 게시 (관리자에겐 신원 확인 가능)'),
    ('AI 챗봇',      '내부회계 관련 24시간 질의응답'),
    ('알림함',       '결재 상신/승인/반려/메모 작성 시 자동 알림'),
])

write_section('8. AI 활용 노하우 (사이트 개발 중 체득)', [
    ('컨텍스트 관리',  '큰 작업은 단계별로 쪼개서 요청 — 한 번에 너무 많은 코드 변경 금지'),
    ('체크포인트',     '주요 작업 후 git tag 로 체크포인트 저장 (예: 완성_26.04.28)'),
    ('스크린샷 활용',   '버그 보고할 때 화면 캡쳐 + 콘솔 로그 함께 전달하면 진단 정확도 ↑'),
    ('자동화 검증',    'puppeteer 로 100회 자동 테스트 → 사람이 못 잡는 회귀 검출'),
    ('실시간 사용자 피드백', 'Tell me / AI 챗봇 / 알림함을 통해 사용자 의견 즉시 수렴'),
])

# ════════════════════════════════════════════
# Sheet 2 — 표 형식 (AI 요약본 양식)
# ════════════════════════════════════════════
ws2 = wb.create_sheet('내용-2(AI요약)')
headers = ['항목', '용도', '기술 및 도구', '세부 사양', '보안 및 정책', '비고']
widths  = [22, 28, 26, 32, 28, 18]
for i, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws2.cell(1, i, h)
    c.font = H_FONT
    c.fill = H_FILL
    c.alignment = CENTER
    c.border = BORDER
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 26

rows = [
    ('백엔드 / DB',          '데이터 저장 및 인증',           'Supabase (PostgreSQL)',     'Pro 플랜, 8GB DB · 100GB 스토리지 (월 $25)', 'RLS · HTTPS · bcrypt 비밀번호 해싱', ''),
    ('프론트엔드 프레임워크', 'UI 화면 구성',                  'React 19.2',                'TypeScript 6.0 + 컴포넌트 기반',         '',                                ''),
    ('CSS 프레임워크',       '디자인 시스템 구현',            'Tailwind CSS 3.4',          '유틸리티 클래스, 토스 톤 brand-700',     '',                                ''),
    ('빌드 도구',            '개발 서버 + 프로덕션 빌드',     'Vite 8.0',                  'HMR(Hot Module Reload) 지원',           '',                                ''),
    ('차트 라이브러리',      'KPI 시각화',                    'Recharts 3.8',              'Bar / Pie / Line / Trend chart',        '',                                ''),
    ('아이콘',               'UI 아이콘',                     'Lucide React 1.7',          '600+ 아이콘',                            '',                                ''),
    ('PDF 렌더링',           'PDF 파일 미리보기',             'pdfjs-dist 5.6',            '브라우저 내 미리보기',                   '',                                ''),
    ('Office 문서 변환',     'DOCX/HWP 미리보기',             'mammoth 1.12',              '브라우저 내 변환',                       '',                                ''),
    ('Excel 다운로드',       '모집단·증빙목록 엑셀 출력',     'xlsx 0.18',                 '동적 import (vendor chunk 분리)',       '',                                ''),
    ('일괄 다운로드',        '증빙 ZIP 묶음 다운로드',        'JSZip 3.10',                'createSignedUrl + fetch + zip',         '',                                ''),
    ('인증',                 '사용자 로그인',                 'Supabase Auth',             '사번 기반, 초기 비밀번호 ty+사번',       '관리자 가입승인 + bcrypt',         ''),
    ('Row Level Security',   '데이터별 권한 차등 적용',       'PostgreSQL RLS Policy',     'Owner/Controller/Admin 별 정책',         '본인 데이터만 조회·수정',          '핵심 보안'),
    ('SECURITY DEFINER RPC', '권한 우회 트랜잭션',            'PostgreSQL 함수',           'set_review_modify_req · notify_*',       'admin 권한 필요한 작업',           ''),
    ('Realtime',             '증빙 카운트 실시간 동기화',     'Supabase Realtime',         'evidence_uploads 테이블 변경 구독',     '',                                ''),
    ('배포',                 '프론트엔드 호스팅',             'Vercel',                    '자동 배포 + CDN + version.json 캐시 정책', 'HTTPS 자동 갱신',                ''),
    ('소스 관리',            '코드 버전 관리',                'GitHub',                    '메인 브랜치 자동 배포',                  '',                                ''),
    ('DART API',             '동양 공시 자동 수집',           'opendart.fss.or.kr',        '일일 갱신 (사업/반기/분기 보고서)',      'API key 환경변수',                ''),
    ('네이버 금융',          '동양 주가 + 뉴스 RSS',          '네이버 모바일 API + RSS',   '실시간 주가 + 카테고리별 뉴스',          '',                                ''),
    ('Google News RSS',      '레미콘/건설/플랜트 뉴스',       'news.google.com RSS',       '4 탭 자동 갱신',                         '',                                ''),
    ('AI 챗봇',              '내부회계 질의응답',             'Anthropic Claude API',      'sonnet/haiku 모델',                      'API key 환경변수',                ''),
    ('Edge TTS',             '교육 영상 내레이션',            'edge-tts (ko-KR-SunHi)',    '한국어 여성 음성',                       '',                                ''),
    ('Puppeteer',            'E2E 자동 테스트 + 영상 녹화',   'puppeteer 24.42 + CDP',     'Page.startScreencast 1080p',             '',                                ''),
    ('python-pptx',          '교육자료 PPTX 자동 생성',       'python-pptx 1.0',           '14~18 슬라이드 자동 빌드',               '',                                ''),
    ('Viewport',             '화면 표시 범위',                '데스크톱 우선',             '1280px ~ 1920px',                        '',                                '모바일 추후'),
    ('Node.js',              '런타임',                        'Node.js',                   '20+ 버전',                                '',                                ''),
    ('npm',                  '패키지 매니저',                 'npm',                       '10+ 버전',                                '',                                ''),
]

for ridx, row in enumerate(rows, 2):
    fill = ALT_FILL if ridx % 2 == 0 else None
    for cidx, val in enumerate(row, 1):
        c = ws2.cell(ridx, cidx, val)
        c.font = CELL_FONT
        c.alignment = LEFT if cidx > 1 else CENTER
        c.border = BORDER
        if fill: c.fill = fill
    ws2.row_dimensions[ridx].height = 32

# ════════════════════════════════════════════
# Sheet 3 — 메뉴 범주별 요약 (개발 노하우 양식)
# ════════════════════════════════════════════
ws3 = wb.create_sheet('내용-3(메뉴별)')

ws3.cell(1, 1, '■ 개발 중 체득한 노하우 (TYIA 사이트)').font = TITLE_FONT
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

tips = [
    '1. 작업은 작은 단위로 나누고, 매 단계마다 빌드 + 검증한다 (한 번에 큰 변경 금지)',
    '2. AI에 컨텍스트를 줄 때는 코드의 핵심 함수 + 에러 메시지 + 스크린샷을 함께 전달한다',
    '3. RLS 정책은 INSERT/SELECT/UPDATE/DELETE 각각 별도로 정의하고 SECURITY DEFINER 함수로 우회 경로 마련',
    '4. UI 변경 후 puppeteer 로 100회 자동 테스트 → 회귀 검출. 사람이 못 보는 race condition 발견됨',
    '5. 사용자가 보고한 버그는 stack trace 그대로 alert 에 노출 → 다음 발생 시 즉시 원인 추적',
    '6. 한글 파일명은 storage 키에서 ASCII 로 변환하고, 원본명은 메타데이터로 별도 저장',
    '7. window.open 으로 PDF 열기는 fetch/realtime channel 점유 위험 — anchor download 만 사용',
    '8. Supabase API 호출은 timeout + retry 래퍼로 감싸서 네트워크 지연 시 사용자 hang 방지',
    '9. 캐시 정책 (Cache-Control: no-store + version.json) 으로 배포 직후 사용자에게 최신 번들 강제',
    '10. git tag 로 주요 마일스톤 체크포인트 저장 → 문제 발생 시 즉시 복구 가능',
]

r = 3
for tip in tips:
    c = ws3.cell(r, 1, tip)
    c.font = CELL_FONT
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
r += 1

# 메뉴 범주별 요약표 (1차강의 양식)
ws3.cell(r, 1, '■ 메뉴 범주별 요약').font = TITLE_FONT
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
headers3 = ['메뉴 범주', '상세 기능 / 페이지', '포함 데이터 및 지표', '주요 기능 설명', '연동 API 및 기술', '사용자 권한 및 보안', '비고']
widths3  = [16, 24, 28, 38, 22, 22, 18]
for i, (h, w) in enumerate(zip(headers3, widths3), 1):
    c = ws3.cell(r, i, h)
    c.font = H_FONT
    c.fill = H_FILL
    c.alignment = CENTER
    c.border = BORDER
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[r].height = 26
r += 1

menus = [
    ('전체 규모', '시스템 통합 관리',
     '통제활동 424건 · 모집단 약 5,000건 · 증빙업로드 12,000+ · 사용자 478명',
     '내부회계관리제도 RCM 전체를 커버. 부서·통제활동·증빙·검토 흐름 일원화.',
     'React + Supabase + Vercel',
     'RLS · is_active 가입승인',
     ''),
    ('대시보드 (홈)', '내 KPI · 검토 상태 · 공지 · 빙고 랭킹',
     '본인 통제활동 진행률 · KPI 점수 · 적립 포인트 · 월말 D-Day',
     '담당자/승인자별 맞춤 카드. 공지·매뉴얼 최신순 노출.',
     'React + Recharts',
     '본인 데이터만 SELECT (RLS)',
     ''),
    ('증빙관리', '증빙 업로드 · 결재상신 · 승인/반려',
     '424 통제활동 · 모집단별 파일 · 결재 이력 · 검토 메모',
     '드래그앤드롭 + 자동 중간저장. 모든 모집단 완료 시 결재상신 활성화.',
     'Supabase Storage + Realtime',
     '담당자: 본인만 / 승인자: 담당 활동만',
     '핵심 메뉴'),
    ('관리자', '사용자 관리 · 통제활동/모집단 업로드 · 모니터링',
     '전사 사용자 478 · 부서/직급 · 권한 · 통제활동 마스터',
     'admin 만 접근. 일괄 업로드(엑셀) + 모니터링 대시보드.',
     'Supabase RPC',
     'role = admin 한정',
     ''),
    ('KPI', '부서별 점수 · 등급(A~F) · 월별 추이',
     '부서별 KPI 92~72점 · 등급 · 완료율 · 누적 활동수',
     '월말 마감 D-Day 와 연계. 자동 등급 산출 및 차트화.',
     'Recharts',
     '전체 공개 (사번 익명)',
     ''),
    ('회사소식', 'DART 공시 + 동양 주가 + 뉴스 4탭',
     '재무제표 (당기/전기/전전기) · 주가 시세 · 카테고리별 뉴스',
     '하루 1회 자동 수집. 출처 표기 (사업/반기/분기 보고서) 명시.',
     'DART API + 네이버 금융 + Google News RSS',
     '전체 공개',
     ''),
    ('빙고', '매일 25문제 · 1줄/오답 = 도전 1회',
     '오답·정답 누적 · 줄 수 · 일 3회 한도 · 월간 랭킹',
     '교육 효과 + 게임화. 1라운드 = 1 도전 카운트 (이중 카운트 방지).',
     'React + LocalStorage',
     '본인 누적만 (DB 저장)',
     ''),
    ('강좌', '내부회계 영상 + 퀴즈 + 익명 질문',
     '영상 진행률 · 수료 여부 · 질문/답변 · 포인트',
     '강좌 별 수료 퀴즈 (정답 1개당 1점). 질문 익명 게시 가능.',
     'pdf.js + Edge TTS (영상)',
     '관리자만 [익명: 홍길동] 표시',
     ''),
    ('웹툰', '내부회계 만화 + 익명 댓글',
     '에피소드 · 댓글 · 좋아요 · 포인트(+5P)',
     '교육 콘텐츠를 만화로. 댓글 익명 옵션 제공.',
     'Supabase + JSON',
     '관리자만 익명 신원 확인',
     ''),
    ('Tell me', '회사 의견·불편·건의 익명 게시',
     '카테고리(응원/불편/개선/건의/쓴소리) · 댓글 · 가림',
     '익명 게시 후 작성자 본인은 수정/삭제 가능. 관리자는 가림 처리.',
     'Supabase',
     '관리자만 익명 신원 확인',
     '핵심 소통'),
    ('AI 챗봇', '내부회계 24시간 질의응답',
     '대화 이력 · 컨텍스트 (사용자 부서/역할)',
     'Anthropic Claude API 호출. 부서별 컨텍스트 주입.',
     'Anthropic API',
     '본인 대화만',
     ''),
    ('알림함', '결재/메모/공지 자동 알림',
     '읽음 여부 · 발신자 · 본문',
     '담당자 → 승인자 → 관리자 흐름 모두 자동 알림.',
     'Supabase + RPC',
     '본인 수신함만 (RLS)',
     ''),
    ('포인트·랭킹', '활동별 포인트 적립',
     '로그인 +10P · 빙고 10P×3 · 강좌 N · 댓글 5P · Tell me 30P',
     '월간 TOP 10 자동 산출. 시상 연계.',
     'Supabase',
     '본인 누적 + 공개 랭킹',
     ''),
]
for menu in menus:
    fill = ALT_FILL if r % 2 == 0 else None
    for cidx, val in enumerate(menu, 1):
        c = ws3.cell(r, cidx, val)
        c.font = CELL_FONT
        c.alignment = LEFT if cidx > 1 else CENTER
        c.border = BORDER
        if fill: c.fill = fill
    ws3.row_dimensions[r].height = 36
    r += 1

# 저장
wb.save(OUT)
print(f"[OK] saved: {OUT}".encode('ascii', 'replace').decode('ascii'))
print(f"     sheets={len(wb.sheetnames)} sheet1_rows={ws.max_row} sheet2_rows={ws2.max_row} sheet3_rows={ws3.max_row}")
